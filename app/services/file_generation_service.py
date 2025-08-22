import os
import uuid
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Optional, Dict, Any
import io
import re
from app.services.file_service import file_service

from app.core.exceptions import FileProcessingException


class FileGenerationService:
    """Excel 파일 생성 서비스"""
    
    def __init__(self):
        self.temp_dir = "temp_files"
        self._ensure_temp_dir()
    
    def _ensure_temp_dir(self):
        """임시 파일 디렉토리 생성"""
        if not os.path.exists(self.temp_dir):
            os.makedirs(self.temp_dir)
    
    def _extract_analysis_content(self, ai_response: str) -> Dict[str, str]:
        """AI 응답에서 분석 내용 추출 - 보고서 형식으로 구조화"""
        content = {
            "제목": "📊 학생 성적 분석 보고서",
            "데이터분석": "",
            "구체적계산": "",
            "결과해석": "",
            "전략적시사점": "",
            "시각화제안": ""
        }
        
        # 제목 추출
        title_patterns = [
            r"📊\s*([^:\n]+)",
            r"🎯\s*([^:\n]+)",
            r"📈\s*([^:\n]+)"
        ]
        
        for pattern in title_patterns:
            match = re.search(pattern, ai_response, re.DOTALL | re.IGNORECASE)
            if match:
                content["제목"] = match.group(1).strip()
                break
        
        # 데이터 분석 방식 추출
        analysis_patterns = [
            r"🎯\s*데이터 분석[:\s]*(.*?)(?=\n\n|\n🔧|\n💡|\n📋|\n🎯|\n📊|\n\[|$)",
            r"데이터\s*분석[:\s]*(.*?)(?=\n\n|\n🔧|\n💡|\n📋|\n🎯|\n📊|\n\[|$)",
            r"분석\s*방식[:\s]*(.*?)(?=\n\n|\n🔧|\n💡|\n📋|\n🎯|\n📊|\n\[|$)"
        ]
        
        for pattern in analysis_patterns:
            match = re.search(pattern, ai_response, re.DOTALL | re.IGNORECASE)
            if match:
                content["데이터분석"] = match.group(1).strip()
                break
        
        # 구체적 계산 (수식) 추출
        calculation_patterns = [
            r"🔧\s*구체적 계산[:\s]*(.*?)(?=\n\n|\n💡|\n📋|\n🎯|\n📊|\n\[|$)",
            r"📋\s*실행 단계[:\s]*(.*?)(?=\n\n|\n💡|\n🔧|\n🎯|\n📊|\n\[|$)",
            r"실행\s*단계[:\s]*(.*?)(?=\n\n|\n💡|\n🔧|\n🎯|\n📊|\n\[|$)",
            r"단계[:\s]*(.*?)(?=\n\n|\n💡|\n🔧|\n🎯|\n📊|\n\[|$)"
        ]
        
        for pattern in calculation_patterns:
            match = re.search(pattern, ai_response, re.DOTALL | re.IGNORECASE)
            if match:
                content["구체적계산"] = match.group(1).strip()
                break
        
        # 결과 해석 추출
        interpretation_patterns = [
            r"💡\s*결과 해석[:\s]*(.*?)(?=\n\n|\n🔧|\n📋|\n🎯|\n📊|\n\[|$)",
            r"결과\s*해석[:\s]*(.*?)(?=\n\n|\n🔧|\n📋|\n🎯|\n📊|\n\[|$)",
            r"해석[:\s]*(.*?)(?=\n\n|\n🔧|\n📋|\n🎯|\n📊|\n\[|$)"
        ]
        
        for pattern in interpretation_patterns:
            match = re.search(pattern, ai_response, re.DOTALL | re.IGNORECASE)
            if match:
                content["결과해석"] = match.group(1).strip()
                break
        
        # 전략적 시사점 추출
        strategy_patterns = [
            r"전략적\s*시사점[:\s]*(.*?)(?=\n\n|\n🔧|\n📋|\n💡|\n🎯|\n📊|\n\[|$)",
            r"시사점[:\s]*(.*?)(?=\n\n|\n🔧|\n📋|\n💡|\n🎯|\n📊|\n\[|$)"
        ]
        
        for pattern in strategy_patterns:
            match = re.search(pattern, ai_response, re.DOTALL | re.IGNORECASE)
            if match:
                content["전략적시사점"] = match.group(1).strip()
                break
        
        # 시각화 제안 추출
        visualization_patterns = [
            r"📊\s*시각화[:\s]*(.*?)(?=\n\n|\n🔧|\n📋|\n💡|\n🎯|\n\[|$)",
            r"시각화[:\s]*(.*?)(?=\n\n|\n🔧|\n📋|\n💡|\n🎯|\n\[|$)"
        ]
        
        for pattern in visualization_patterns:
            match = re.search(pattern, ai_response, re.DOTALL | re.IGNORECASE)
            if match:
                content["시각화제안"] = match.group(1).strip()
                break
        
        # 내용이 없으면 전체 응답에서 주요 내용 추출
        if not any([content["데이터분석"], content["구체적계산"], content["결과해석"]]):
            lines = ai_response.split('\n')
            important_lines = []
            for line in lines:
                line = line.strip()
                if line and not line.startswith('[') and len(line) > 10:
                    important_lines.append(line)
            
            if important_lines:
                content["데이터분석"] = '\n'.join(important_lines[:3])
                if len(important_lines) > 3:
                    content["결과해석"] = '\n'.join(important_lines[3:6])
        
        return content
    
    def generate_analysis_file(self, session_id: str, ai_response: str, original_file_content: bytes, selected_sheet: str = None, original_filename: str = None) -> Dict[str, Any]:
        """AI 응답을 기반으로 간단한 Excel 파일 생성 - 사용자가 요청한 작업만 수행"""
        try:
            # 파일 ID 생성
            file_id = str(uuid.uuid4())[:8]
            file_path = os.path.join(self.temp_dir, f"{file_id}.xlsx")
            
            # 원본 파일 분석
            actual_filename = original_filename or "original_file.xlsx"
            analysis_result = file_service.analyze_excel_file(original_file_content, actual_filename)
            
            # 새로운 워크북 생성
            wb = Workbook()
            
            # 사용자가 요청한 작업만 수행 - 간단한 데이터 처리
            if analysis_result.sheets:
                if selected_sheet and selected_sheet in analysis_result.sheets:
                    original_sheet_name = selected_sheet
                else:
                    original_sheet_name = analysis_result.sheets[0]
                
                # 원본 데이터 읽기
                df_original = pd.read_excel(io.BytesIO(original_file_content), sheet_name=original_sheet_name)
                
                # AI 응답에서 빈 셀 채우기 등의 간단한 작업만 수행
                # 복잡한 분석은 제거
                
                ws_original = wb.active
                ws_original.title = f"처리된_{original_sheet_name}"
                
                # 처리된 데이터 쓰기
                for r in dataframe_to_rows(df_original, index=False, header=True):
                    ws_original.append(r)
                
                # 헤더 스타일 적용
                for cell in ws_original[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            
            # 열 너비 자동 조정
            for ws in wb.worksheets:
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # 파일 저장
            wb.save(file_path)
            
            return {
                'file_id': file_id,
                'download_url': f'/api/chat/download/{file_id}',
                'message': '요청하신 작업이 완료된 파일이 생성되었습니다.'
            }
            
        except Exception as e:
            print(f"❌ 파일 생성 중 오류: {e}")
            raise FileProcessingException(f"파일 생성 실패: {str(e)}")
    
    def get_file_path(self, file_id: str) -> Optional[str]:
        """파일 ID로 파일 경로 반환"""
        file_path = os.path.join(self.temp_dir, f"{file_id}.xlsx")
        if os.path.exists(file_path):
            return file_path
        return None

    def _create_advanced_statistics_sheet(self, ws, analysis: Dict):
        """고급 통계분석 시트 생성"""
        # 제목
        ws.merge_cells('A1:D1')
        title = ws['A1']
        title.value = "📊 고급 통계분석 결과"
        title.font = Font(size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 3
        
        # 기술통계
        if 'statistics' in analysis and 'descriptive' in analysis['statistics']:
            ws.merge_cells(f'A{row}:D{row}')
            section_title = ws[f'A{row}']
            section_title.value = "📈 기술통계"
            section_title.font = Font(size=14, bold=True, color="FFFFFF")
            section_title.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            row += 1
            
            # 기술통계 테이블 작성
            desc_stats = analysis['statistics']['descriptive']
            for col in desc_stats.columns:
                ws.cell(row=row, column=1, value=col).font = Font(bold=True)
                for i, stat in enumerate(desc_stats.index):
                    ws.cell(row=row+i+1, column=1, value=stat)
                    ws.cell(row=row+i+1, column=2, value=f"{desc_stats.loc[stat, col]:.2f}")
                row += len(desc_stats) + 2
        
        # 왜도 및 첨도
        if 'statistics' in analysis and 'skewness' in analysis['statistics']:
            ws.merge_cells(f'A{row}:D{row}')
            section_title = ws[f'A{row}']
            section_title.value = "📊 분포 특성"
            section_title.font = Font(size=14, bold=True, color="FFFFFF")
            section_title.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
            row += 1
            
            skew_data = analysis['statistics']['skewness']
            kurt_data = analysis['statistics']['kurtosis']
            
            for col in skew_data.index:
                ws.cell(row=row, column=1, value=col).font = Font(bold=True)
                ws.cell(row=row, column=2, value=f"왜도: {skew_data[col]:.3f}")
                ws.cell(row=row, column=3, value=f"첨도: {kurt_data[col]:.3f}")
                row += 1
            row += 1
    
    def _create_advanced_insights_sheet(self, ws, analysis: Dict):
        """고급 데이터 인사이트 시트 생성"""
        # 제목
        ws.merge_cells('A1:D1')
        title = ws['A1']
        title.value = "💡 데이터 인사이트"
        title.font = Font(size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 3
        
        if 'insights' in analysis:
            for insight in analysis['insights']:
                ws.merge_cells(f'A{row}:D{row}')
                insight_cell = ws[f'A{row}']
                insight_cell.value = insight
                insight_cell.font = Font(size=11)
                insight_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                row += 1
        
        # 추가 인사이트
        if 'patterns' in analysis:
            # 상관관계 인사이트
            if 'correlations' in analysis['patterns'] and analysis['patterns']['correlations']:
                ws.merge_cells(f'A{row}:D{row}')
                section_title = ws[f'A{row}']
                section_title.value = "🔗 상관관계 분석"
                section_title.font = Font(size=12, bold=True, color="FFFFFF")
                section_title.fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
                row += 1
                
                for corr in analysis['patterns']['correlations']:
                    insight_text = f"{corr['col1']} ↔ {corr['col2']}: {corr['correlation']:.3f}"
                    ws.merge_cells(f'A{row}:D{row}')
                    ws[f'A{row}'].value = insight_text
                    row += 1
                row += 1
    
    def _create_data_quality_sheet(self, ws, analysis: Dict):
        """데이터 품질 진단 시트 생성"""
        # 제목
        ws.merge_cells('A1:D1')
        title = ws['A1']
        title.value = "🔍 데이터 품질 진단"
        title.font = Font(size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 3
        
        # 결측치 정보
        if 'patterns' in analysis and 'missing_data' in analysis['patterns']:
            ws.merge_cells(f'A{row}:D{row}')
            section_title = ws[f'A{row}']
            section_title.value = "❓ 결측치 분석"
            section_title.font = Font(size=14, bold=True, color="FFFFFF")
            section_title.fill = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
            row += 1
            
            for col, info in analysis['patterns']['missing_data'].items():
                ws.cell(row=row, column=1, value=col).font = Font(bold=True)
                ws.cell(row=row, column=2, value=f"{info['count']}개 ({info['percentage']:.1f}%)")
                row += 1
        
        row += 2
        
        # 이상치 정보
        if 'patterns' in analysis and 'outliers' in analysis['patterns']:
            ws.merge_cells(f'A{row}:D{row}')
            section_title = ws[f'A{row}']
            section_title.value = "⚠️ 이상치 분석"
            section_title.font = Font(size=14, bold=True, color="FFFFFF")
            section_title.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
            row += 1
            
            for col, info in analysis['patterns']['outliers'].items():
                ws.cell(row=row, column=1, value=col).font = Font(bold=True)
                ws.cell(row=row, column=2, value=f"{info['count']}개 ({info['percentage']:.1f}%)")
                row += 1
    
    def _create_optimization_recommendations_sheet(self, ws, analysis: Dict):
        """최적화 추천 시트 생성"""
        # 제목
        ws.merge_cells('A1:D1')
        title = ws['A1']
        title.value = "🎯 최적화 추천"
        title.font = Font(size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 3
        
        if 'recommendations' in analysis:
            # 차트 추천
            if 'charts' in analysis['recommendations']:
                ws.merge_cells(f'A{row}:D{row}')
                section_title = ws[f'A{row}']
                section_title.value = "📊 추천 차트"
                section_title.font = Font(size=14, bold=True, color="FFFFFF")
                section_title.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                row += 1
                
                for chart_rec in analysis['recommendations']['charts']:
                    ws.merge_cells(f'A{row}:D{row}')
                    chart_cell = ws[f'A{row}']
                    chart_cell.value = f"📈 {chart_rec['title']}: {chart_rec['description']}"
                    chart_cell.font = Font(size=11)
                    chart_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    row += 1
                row += 1
            
            # 조건부 서식 추천
            if 'formats' in analysis['recommendations']:
                ws.merge_cells(f'A{row}:D{row}')
                section_title = ws[f'A{row}']
                section_title.value = "🎨 조건부 서식"
                section_title.font = Font(size=14, bold=True, color="FFFFFF")
                section_title.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
                row += 1
                
                for format_rec in analysis['recommendations']['formats']:
                    ws.merge_cells(f'A{row}:D{row}')
                    format_cell = ws[f'A{row}']
                    format_cell.value = f"🎨 {format_rec['column']}: {format_rec['description']}"
                    format_cell.font = Font(size=11)
                    format_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    row += 1
                row += 1
            
            # 분석 추천
            if 'analysis' in analysis['recommendations']:
                ws.merge_cells(f'A{row}:D{row}')
                section_title = ws[f'A{row}']
                section_title.value = "🔬 추가 분석"
                section_title.font = Font(size=14, bold=True, color="FFFFFF")
                section_title.fill = PatternFill(start_color="9C27B0", end_color="9C27B0", fill_type="solid")
                row += 1
                
                for analysis_rec in analysis['recommendations']['analysis']:
                    ws.merge_cells(f'A{row}:D{row}')
                    analysis_cell = ws[f'A{row}']
                    analysis_cell.value = f"🔬 {analysis_rec['description']}"
                    analysis_cell.font = Font(size=11)
                    analysis_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    row += 1
                row += 1
            
            # 액션 추천
            if 'actions' in analysis['recommendations']:
                ws.merge_cells(f'A{row}:D{row}')
                section_title = ws[f'A{row}']
                section_title.value = "⚡ 권장 액션"
                section_title.font = Font(size=14, bold=True, color="FFFFFF")
                section_title.fill = PatternFill(start_color="FF9800", end_color="FF9800", fill_type="solid")
                row += 1
                
                for action_rec in analysis['recommendations']['actions']:
                    ws.merge_cells(f'A{row}:D{row}')
                    action_cell = ws[f'A{row}']
                    action_cell.value = f"⚡ {action_rec['description']}"
                    action_cell.font = Font(size=11)
                    action_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    row += 1
    
    def _create_visualization_sheet(self, ws, df: pd.DataFrame, analysis: Dict):
        """시각화 시트 생성 - 실제 차트 및 피벗테이블 포함"""
        # 제목
        ws.merge_cells('A1:D1')
        title = ws['A1']
        title.value = "📊 데이터 시각화"
        title.font = Font(size=16, bold=True, color="FFFFFF")
        title.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title.alignment = Alignment(horizontal="center", vertical="center")
        
        row = 3
        
        # 추천 차트 생성
        if 'recommendations' in analysis and 'charts' in analysis['recommendations']:
            for i, chart_rec in enumerate(analysis['recommendations']['charts']):
                # 차트 설명
                ws.merge_cells(f'A{row}:D{row}')
                chart_title = ws[f'A{row}']
                chart_title.value = f"📈 {chart_rec['title']}"
                chart_title.font = Font(size=12, bold=True)
                row += 1
                
                # 차트 데이터를 시트에 작성
                chart_data_row = row
                try:
                    if chart_rec['type'] == 'bar' and 'data' in chart_rec and isinstance(chart_rec['data'], dict):
                        x_col = chart_rec['data']['x']
                        y_col = chart_rec['data']['y']
                        
                        # 범주별 평균 계산
                        grouped_data = df.groupby(x_col)[y_col].mean().reset_index()
                        
                        # 데이터를 시트에 작성
                        ws.cell(row=row, column=1, value=x_col).font = Font(bold=True)
                        ws.cell(row=row, column=2, value=y_col).font = Font(bold=True)
                        row += 1
                        
                        for _, data_row in grouped_data.iterrows():
                            ws.cell(row=row, column=1, value=data_row[x_col])
                            ws.cell(row=row, column=2, value=data_row[y_col])
                            row += 1
                        
                        # 차트 생성
                        from openpyxl.chart import BarChart, Reference
                        chart = BarChart()
                        chart.title = chart_rec['title']
                        chart.x_axis.title = x_col
                        chart.y_axis.title = y_col
                        
                        # 시트의 데이터를 참조
                        data = Reference(ws, min_col=2, min_row=chart_data_row, max_row=row-1)
                        cats = Reference(ws, min_col=1, min_row=chart_data_row+1, max_row=row-1)
                        
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(cats)
                        
                        # 차트 위치 설정
                        chart.anchor = f'E{chart_data_row}'
                        ws.add_chart(chart)
                        row += 15  # 차트 공간
                        
                    elif chart_rec['type'] == 'scatter' and 'data' in chart_rec and len(chart_rec['data']) >= 2:
                        x_col = chart_rec['data'][0]
                        y_col = chart_rec['data'][1]
                        
                        # 데이터를 시트에 작성
                        ws.cell(row=row, column=1, value=x_col).font = Font(bold=True)
                        ws.cell(row=row, column=2, value=y_col).font = Font(bold=True)
                        row += 1
                        
                        for _, data_row in df.iterrows():
                            ws.cell(row=row, column=1, value=data_row[x_col])
                            ws.cell(row=row, column=2, value=data_row[y_col])
                            row += 1
                        
                        # 차트 생성
                        from openpyxl.chart import ScatterChart, Reference
                        chart = ScatterChart()
                        chart.title = chart_rec['title']
                        chart.x_axis.title = x_col
                        chart.y_axis.title = y_col
                        
                        # 시트의 데이터를 참조
                        data = Reference(ws, min_col=1, min_row=chart_data_row, max_col=2, max_row=row-1)
                        chart.add_data(data, titles_from_data=True)
                        
                        # 차트 위치 설정
                        chart.anchor = f'E{chart_data_row}'
                        ws.add_chart(chart)
                        row += 15  # 차트 공간
                        
                    elif chart_rec['type'] == 'histogram' and 'data' in chart_rec:
                        col_name = chart_rec['data']
                        
                        # 히스토그램을 위한 구간별 데이터 생성
                        data_values = df[col_name].dropna()
                        hist_data, bin_edges = pd.cut(data_values, bins=5, retbins=True)
                        hist_counts = hist_data.value_counts().sort_index()
                        
                        # 데이터를 시트에 작성
                        ws.cell(row=row, column=1, value="구간").font = Font(bold=True)
                        ws.cell(row=row, column=2, value="빈도").font = Font(bold=True)
                        row += 1
                        
                        for interval, count in hist_counts.items():
                            ws.cell(row=row, column=1, value=str(interval))
                            ws.cell(row=row, column=2, value=count)
                            row += 1
                        
                        # 막대 차트로 히스토그램 표현
                        from openpyxl.chart import BarChart, Reference
                        chart = BarChart()
                        chart.title = chart_rec['title']
                        chart.x_axis.title = "구간"
                        chart.y_axis.title = "빈도"
                        
                        # 시트의 데이터를 참조
                        data = Reference(ws, min_col=2, min_row=chart_data_row, max_row=row-1)
                        cats = Reference(ws, min_col=1, min_row=chart_data_row+1, max_row=row-1)
                        
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(cats)
                        
                        # 차트 위치 설정
                        chart.anchor = f'E{chart_data_row}'
                        ws.add_chart(chart)
                        row += 15  # 차트 공간
                        
                except Exception as e:
                    print(f"차트 생성 실패: {e}")
                    row += 5
                
                row += 2  # 차트 간 간격
        
        # 피벗테이블 생성
        row += 2
        ws.merge_cells(f'A{row}:D{row}')
        pivot_title = ws[f'A{row}']
        pivot_title.value = "📋 피벗테이블 분석"
        pivot_title.font = Font(size=14, bold=True, color="FFFFFF")
        pivot_title.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        row += 1
        
        # 피벗테이블 데이터 준비
        pivot_data = self._create_pivot_table_data(df, analysis)
        if pivot_data:
            # 피벗테이블 데이터를 시트에 작성
            for r_idx, row_data in enumerate(pivot_data):
                for c_idx, cell_value in enumerate(row_data):
                    ws.cell(row=row+r_idx, column=c_idx+1, value=cell_value)
            row += len(pivot_data) + 2
    
    def _create_chart(self, df: pd.DataFrame, chart_rec: Dict, chart_name: str):
        """실제 차트 생성"""
        try:
            from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference
            
            chart_type = chart_rec['type']
            
            if chart_type == 'bar':
                chart = BarChart()
                chart.title = chart_rec['title']
                chart.x_axis.title = "범주"
                chart.y_axis.title = "값"
                
                # 데이터 준비
                if 'data' in chart_rec and isinstance(chart_rec['data'], dict):
                    x_col = chart_rec['data']['x']
                    y_col = chart_rec['data']['y']
                    
                    # 범주별 평균 계산
                    grouped_data = df.groupby(x_col)[y_col].mean().reset_index()
                    
                    # 차트 데이터 추가 - 올바른 참조 방식으로 수정
                    data = Reference(grouped_data, min_col=grouped_data.columns.get_loc(y_col)+1, 
                                   min_row=1, max_row=len(grouped_data)+1)
                    cats = Reference(grouped_data, min_col=grouped_data.columns.get_loc(x_col)+1, 
                                   min_row=2, max_row=len(grouped_data)+1)
                    
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
                
            elif chart_type == 'scatter':
                chart = ScatterChart()
                chart.title = chart_rec['title']
                chart.x_axis.title = chart_rec['data'][0]
                chart.y_axis.title = chart_rec['data'][1]
                
                # 산점도 데이터 - 올바른 참조 방식으로 수정
                data = Reference(df, min_col=df.columns.get_loc(chart_rec['data'][0])+1, 
                               min_row=1, max_row=len(df)+1, max_col=df.columns.get_loc(chart_rec['data'][1])+1)
                chart.add_data(data, titles_from_data=True)
                
            elif chart_type == 'line':
                chart = LineChart()
                chart.title = chart_rec['title']
                
                # 시계열 데이터 (날짜 컬럼이 있는 경우)
                date_cols = [col for col in df.columns if self._is_date_column(df[col])]
                if date_cols and len(chart_rec['data']) > 0:
                    numeric_col = chart_rec['data'][0]
                    date_col = date_cols[0]
                    
                    # 날짜별 평균 계산
                    df_temp = df.copy()
                    df_temp[date_col] = pd.to_datetime(df_temp[date_col])
                    grouped_data = df_temp.groupby(df_temp[date_col].dt.date)[numeric_col].mean().reset_index()
                    
                    data = Reference(grouped_data, min_col=grouped_data.columns.get_loc(numeric_col)+1, 
                                   min_row=1, max_row=len(grouped_data)+1)
                    cats = Reference(grouped_data, min_col=grouped_data.columns.get_loc(date_col)+1, 
                                   min_row=2, max_row=len(grouped_data)+1)
                    
                    chart.add_data(data, titles_from_data=True)
                    chart.set_categories(cats)
            
            else:
                return None
            
            return chart
            
        except Exception as e:
            print(f"차트 생성 오류: {e}")
            return None
    
    def _create_pivot_table_data(self, df: pd.DataFrame, analysis: Dict):
        """피벗테이블 데이터 생성"""
        try:
            pivot_data = []
            
            # 헤더
            pivot_data.append(["피벗테이블 분석 결과"])
            pivot_data.append([])
            
            # 데이터 타입별 분석
            numeric_cols = [col for col, dtype in analysis['data_type'].items() if dtype == 'numeric']
            categorical_cols = [col for col, dtype in analysis['data_type'].items() if dtype == 'categorical']
            
            if numeric_cols and categorical_cols:
                # 범주별 수치 요약
                pivot_data.append(["범주별 수치 요약"])
                pivot_data.append(["범주", "평균", "표준편차", "최소값", "최대값", "개수"])
                
                for cat_col in categorical_cols[:2]:  # 최대 2개 범주만
                    for num_col in numeric_cols[:2]:  # 최대 2개 수치만
                        grouped = df.groupby(cat_col)[num_col].agg(['mean', 'std', 'min', 'max', 'count']).reset_index()
                        
                        for _, row in grouped.iterrows():
                            pivot_data.append([
                                f"{cat_col}: {row[cat_col]}",
                                f"{row['mean']:.2f}",
                                f"{row['std']:.2f}",
                                f"{row['min']:.2f}",
                                f"{row['max']:.2f}",
                                f"{int(row['count'])}"
                            ])
                
                pivot_data.append([])
            
            # 상관관계 요약
            if len(numeric_cols) > 1:
                pivot_data.append(["상관관계 분석"])
                pivot_data.append(["변수1", "변수2", "상관계수", "해석"])
                
                corr_matrix = df[numeric_cols].corr()
                for i in range(len(corr_matrix.columns)):
                    for j in range(i+1, len(corr_matrix.columns)):
                        corr_value = corr_matrix.iloc[i, j]
                        strength = "강함" if abs(corr_value) > 0.7 else "보통" if abs(corr_value) > 0.3 else "약함"
                        direction = "양의" if corr_value > 0 else "음의"
                        
                        pivot_data.append([
                            corr_matrix.columns[i],
                            corr_matrix.columns[j],
                            f"{corr_value:.3f}",
                            f"{strength} {direction} 상관관계"
                        ])
            
            return pivot_data
            
        except Exception as e:
            print(f"피벗테이블 데이터 생성 오류: {e}")
            return [["피벗테이블 생성 중 오류가 발생했습니다."]]
    
    def _is_date_column(self, series: pd.Series) -> bool:
        """날짜 형식 컬럼인지 확인"""
        try:
            pd.to_datetime(series, errors='raise')
            return True
        except:
            return False


# 전역 인스턴스
file_generation_service = FileGenerationService()
